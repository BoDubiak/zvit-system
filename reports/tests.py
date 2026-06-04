import tempfile
from io import BytesIO
from zipfile import ZipFile

from django.contrib.auth import get_user_model
from django.core import mail
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.test import TestCase, override_settings
from django.urls import reverse
from openpyxl import load_workbook

from .constants import DEFAULT_REPORT_FORMS
from .models import (
    EmailNotification,
    ExpectedReport,
    Organization,
    OrganizationUser,
    ReportForm,
    ReportingPeriod,
    ReportStatusLog,
)
from .services import normalized_report_filename, parse_financial_xml, validate_uploaded_report


def xml_file(edrpou="20809229", year=2025, month="06", schema="S0100115", name="report.xml"):
    content = f"""<?xml version="1.0" encoding="windows-1251"?>
<DECLAR xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:noNamespaceSchemaLocation="{schema}.xsd">
  <DECLARBODY>
    <TIN>{edrpou}</TIN>
    <PERIOD_YEAR>{year}</PERIOD_YEAR>
    <PERIOD_MONTH>{month}</PERIOD_MONTH>
  </DECLARBODY>
</DECLAR>""".encode("cp1251")
    return SimpleUploadedFile(name, content, content_type="text/xml")


@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class FinancialReportTests(TestCase):
    def setUp(self):
        for item in DEFAULT_REPORT_FORMS:
            ReportForm.objects.create(**item)
        self.organization = Organization.objects.create(name="КП Тест", edrpou="20809229", report_type=Organization.ReportType.FULL)
        self.period = ReportingPeriod.objects.create(year=2025, quarter=ReportingPeriod.Quarter.Q2)
        self.form = ReportForm.objects.get(xml_schema="S0100115")
        self.expected_report = ExpectedReport.objects.create(
            organization=self.organization,
            period=self.period,
            form=self.form,
        )

    def create_staff_user(self):
        return get_user_model().objects.create_user(username="admin", password="pass", is_staff=True)

    def login_staff(self):
        user = self.create_staff_user()
        self.client.force_login(user)
        return user

    def login_organization_admin(self, organization=None):
        user = get_user_model().objects.create_user(username="org-admin", password="pass")
        OrganizationUser.objects.create(
            user=user,
            organization=organization or self.organization,
            role=OrganizationUser.Role.ADMIN,
        )
        self.client.force_login(user)
        return user

    def test_parse_financial_xml(self):
        parsed = parse_financial_xml(xml_file())

        self.assertEqual(parsed["edrpou"], "20809229")
        self.assertEqual(parsed["year"], 2025)
        self.assertEqual(parsed["quarter"], "Q2")
        self.assertEqual(parsed["xml_schema"], "S0100115")

    def test_quarter_detection(self):
        cases = [("03", "Q1"), ("3", "Q1"), ("06", "Q2"), ("6", "Q2"), ("09", "Q3"), ("9", "Q3"), ("12", "Q4")]
        for month, quarter in cases:
            with self.subTest(month=month):
                self.assertEqual(parse_financial_xml(xml_file(month=month))["quarter"], quarter)

    def test_successful_validation(self):
        ok, message = validate_uploaded_report(self.expected_report, xml_file())
        self.expected_report.refresh_from_db()

        self.assertTrue(ok)
        self.assertEqual(message, "Файл успішно перевірено")
        self.assertEqual(self.expected_report.status, ExpectedReport.Status.UPLOADED)
        self.assertEqual(self.expected_report.normalized_filename, "20809229-2025-Q2.XML")
        self.assertEqual(self.expected_report.upload_logs.count(), 1)
        self.assertEqual(
            self.expected_report.uploaded_file.name,
            "reports/J0900108/2025/Q2/20809229/20809229-2025-Q2.XML",
        )
        self.assertTrue(
            self.expected_report.upload_logs.get().file.name.startswith(
                "upload_logs/J0900108/2025/Q2/20809229/"
            )
        )

    def test_validation_error_when_edrpou_mismatch(self):
        ok, message = validate_uploaded_report(self.expected_report, xml_file(edrpou="01984300"))
        self.expected_report.refresh_from_db()

        self.assertFalse(ok)
        self.assertIn("Очікується ЄДРПОУ 20809229", message)
        self.assertEqual(self.expected_report.status, ExpectedReport.Status.REJECTED)

    def test_repeated_upload_overwrites_current_file_at_stable_path(self):
        validate_uploaded_report(self.expected_report, xml_file())
        validate_uploaded_report(self.expected_report, xml_file())
        self.expected_report.refresh_from_db()

        self.assertEqual(
            self.expected_report.uploaded_file.name,
            "reports/J0900108/2025/Q2/20809229/20809229-2025-Q2.XML",
        )
        self.assertEqual(self.expected_report.upload_logs.count(), 2)

    def test_validation_error_when_form_mismatch(self):
        ok, message = validate_uploaded_report(self.expected_report, xml_file(schema="S0100215"))
        self.expected_report.refresh_from_db()

        self.assertFalse(ok)
        self.assertIn("Очікується форма S0100115", message)
        self.assertEqual(self.expected_report.status, ExpectedReport.Status.REJECTED)

    def test_generate_expected_reports_for_small_and_full_required_forms(self):
        Organization.objects.create(name="КП Мале", edrpou="12345678", report_type=Organization.ReportType.SMALL)
        call_command("generate_expected_reports", year=2025, quarter="Q2", verbosity=0)

        full_reports = ExpectedReport.objects.filter(organization=self.organization, period=self.period)
        small_reports = ExpectedReport.objects.filter(organization__edrpou="12345678", period=self.period)

        self.assertEqual(full_reports.count(), 2)
        self.assertCountEqual(
            full_reports.values_list("form__xml_schema", flat=True),
            ["S0100115", "S0100215"],
        )
        self.assertEqual(small_reports.count(), 1)
        self.assertEqual(small_reports.get().form.xml_schema, "S0110014")

    def test_generate_expected_reports_queues_grouped_email_notification(self):
        self.organization.contact_email = "contact@example.com"
        self.organization.save(update_fields=["contact_email"])
        user = get_user_model().objects.create_user(
            username="notified-user",
            password="pass",
            email="user@example.com",
        )
        OrganizationUser.objects.create(user=user, organization=self.organization)

        with self.captureOnCommitCallbacks(execute=True):
            call_command("generate_expected_reports", year=2026, quarter="Q1", verbosity=0)

        notification = EmailNotification.objects.get()
        self.assertEqual(notification.organization, self.organization)
        self.assertEqual(notification.period.year, 2026)
        self.assertEqual(notification.period.quarter, "Q1")
        self.assertEqual(notification.status, EmailNotification.Status.PENDING)
        self.assertCountEqual(notification.recipients, ["contact@example.com", "user@example.com"])
        self.assertIn("J0900108", notification.body)
        self.assertIn("J0900207", notification.body)

    @override_settings(
        EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
        DEFAULT_FROM_EMAIL="system@example.com",
    )
    def test_send_email_notifications_command_sends_pending_notification(self):
        notification = EmailNotification.objects.create(
            organization=self.organization,
            period=self.period,
            notification_type=EmailNotification.Type.EXPECTED_REPORTS_CREATED,
            recipients=["user@example.com"],
            subject="Test notification",
            body="Notification body",
        )

        call_command("send_email_notifications", verbosity=0)

        notification.refresh_from_db()
        self.assertEqual(notification.status, EmailNotification.Status.SENT)
        self.assertEqual(notification.attempts, 1)
        self.assertIsNotNone(notification.sent_at)
        self.assertEqual(len(mail.outbox), 1)
        self.assertEqual(mail.outbox[0].subject, "Test notification")
        self.assertEqual(mail.outbox[0].to, ["user@example.com"])

    def test_generate_expected_reports_can_include_optional_full_forms(self):
        call_command("generate_expected_reports", year=2025, quarter="Q2", include_optional=True, verbosity=0)

        full_reports = ExpectedReport.objects.filter(organization=self.organization, period=self.period)

        self.assertEqual(full_reports.count(), 5)
        self.assertCountEqual(
            full_reports.values_list("form__xml_schema", flat=True),
            ["S0100115", "S0100215", "S0100311", "S0104010", "S0105009"],
        )

    def test_normalized_filename(self):
        self.assertEqual(normalized_report_filename(self.expected_report), "20809229-2025-Q2.XML")

    def test_staff_can_generate_expected_reports_from_site(self):
        self.login_staff()
        Organization.objects.create(name="КП Мале", edrpou="12345678", report_type=Organization.ReportType.SMALL)

        response = self.client.post(
            reverse("generate_expected_reports"),
            {"year": 2026, "quarter": "Q1"},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        period = ReportingPeriod.objects.get(year=2026, quarter="Q1")
        self.assertEqual(ExpectedReport.objects.filter(period=period, organization=self.organization).count(), 2)
        self.assertEqual(ExpectedReport.objects.filter(period=period, organization__edrpou="12345678").count(), 1)

    def test_staff_can_generate_selected_forms_for_selected_organization(self):
        self.login_staff()
        other = Organization.objects.create(name="КП Інше", edrpou="12345678", report_type=Organization.ReportType.FULL)
        optional_form = ReportForm.objects.get(code="J0900904")

        response = self.client.post(
            reverse("generate_expected_reports"),
            {
                "year": 2026,
                "quarter": "Q2",
                "organizations": [other.id],
                "report_forms": [optional_form.id],
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        period = ReportingPeriod.objects.get(year=2026, quarter="Q2")
        self.assertFalse(ExpectedReport.objects.filter(period=period, organization=self.organization).exists())
        report = ExpectedReport.objects.get(period=period, organization=other)
        self.assertEqual(report.form.code, "J0900904")

    def test_admin_dashboard_can_filter_uploaded_reports(self):
        self.login_staff()
        validate_uploaded_report(self.expected_report, xml_file())

        response = self.client.get(reverse("admin_dashboard"), {"status": "uploaded"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "20809229")
        self.assertContains(response, "Завантажено")

    def test_staff_without_organization_is_redirected_to_admin_dashboard(self):
        self.login_staff()

        response = self.client.get(reverse("company_reports"))

        self.assertRedirects(response, reverse("admin_dashboard"))

    def test_staff_without_organization_does_not_see_company_dashboard_link(self):
        self.login_staff()

        response = self.client.get(reverse("admin_dashboard"))

        self.assertNotContains(response, f'href="{reverse("company_reports")}"')

    def test_admin_dashboard_links_excel_export_before_period_filter(self):
        self.login_staff()

        response = self.client.get(reverse("admin_dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, f'href="{reverse("export_control_report")}"')

    def test_staff_can_export_excel_for_all_periods_from_site(self):
        self.login_staff()
        validate_uploaded_report(self.expected_report, xml_file())

        response = self.client.get(reverse("export_control_report"))

        self.assertEqual(response.status_code, 200)
        self.assertIn("control_report_all_periods.xlsx", response["Content-Disposition"])
        content = b"".join(response.streaming_content)
        workbook = load_workbook(BytesIO(content), read_only=True)
        received_rows = list(workbook["received_files"].iter_rows(values_only=True))
        self.assertIn(
            (
                self.organization.name,
                self.organization.edrpou,
                self.period.year,
                str(self.period.quarter),
                self.form.code,
                self.form.xml_schema,
                str(ExpectedReport.Status.UPLOADED),
                "report.xml",
                "20809229-2025-Q2.XML",
                self.expected_report.uploaded_at.isoformat(),
                None,
            ),
            received_rows,
        )

    def test_staff_can_accept_report_from_dashboard(self):
        user = self.login_staff()
        validate_uploaded_report(self.expected_report, xml_file())

        response = self.client.post(reverse("accept_expected_report", args=[self.expected_report.id]), follow=True)
        self.expected_report.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.expected_report.status, ExpectedReport.Status.ACCEPTED)
        self.assertEqual(self.expected_report.validation_message, "Звіт прийнято адміністратором")
        status_log = ReportStatusLog.objects.get(expected_report=self.expected_report)
        self.assertEqual(status_log.changed_by, user)
        self.assertEqual(status_log.old_status, ExpectedReport.Status.UPLOADED)
        self.assertEqual(status_log.new_status, ExpectedReport.Status.ACCEPTED)

    def test_staff_cannot_accept_pending_report_without_file(self):
        self.login_staff()

        response = self.client.post(reverse("accept_expected_report", args=[self.expected_report.id]), follow=True)
        self.expected_report.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.expected_report.status, ExpectedReport.Status.PENDING)
        self.assertFalse(ReportStatusLog.objects.filter(expected_report=self.expected_report).exists())
        self.assertContains(response, "Прийняти можна тільки завантажений звіт із файлом.")

    def test_admin_dashboard_hides_accept_button_for_pending_report(self):
        self.login_staff()

        response = self.client.get(reverse("admin_dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, reverse("accept_expected_report", args=[self.expected_report.id]))

    def test_staff_can_reject_report_from_dashboard(self):
        user = self.login_staff()
        validate_uploaded_report(self.expected_report, xml_file())

        response = self.client.post(
            reverse("reject_expected_report", args=[self.expected_report.id]),
            {"reason": "Потрібно завантажити правильний файл"},
            follow=True,
        )
        self.expected_report.refresh_from_db()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.expected_report.status, ExpectedReport.Status.REJECTED)
        self.assertEqual(self.expected_report.validation_message, "Потрібно завантажити правильний файл")
        status_log = ReportStatusLog.objects.get(expected_report=self.expected_report)
        self.assertEqual(status_log.changed_by, user)
        self.assertEqual(status_log.old_status, ExpectedReport.Status.UPLOADED)
        self.assertEqual(status_log.new_status, ExpectedReport.Status.REJECTED)
        self.assertEqual(status_log.comment, "Потрібно завантажити правильний файл")

    def test_staff_can_export_zip_archives_from_site(self):
        self.login_staff()
        validate_uploaded_report(self.expected_report, xml_file())

        response = self.client.get(reverse("export_archives"), {"period": self.period.id})

        self.assertEqual(response.status_code, 200)
        content = b"".join(response.streaming_content)
        with ZipFile(BytesIO(content)) as bundle:
            self.assertIn("J0900108.zip", bundle.namelist())
            with bundle.open("J0900108.zip") as form_zip_file:
                with ZipFile(BytesIO(form_zip_file.read())) as form_zip:
                    self.assertIn("20809229-2025-Q2.XML", form_zip.namelist())

    def test_staff_can_export_zip_archives_for_all_periods_from_site(self):
        self.login_staff()
        validate_uploaded_report(self.expected_report, xml_file())
        other_period = ReportingPeriod.objects.create(year=2025, quarter=ReportingPeriod.Quarter.Q1)
        other_report = ExpectedReport.objects.create(
            organization=self.organization,
            period=other_period,
            form=self.form,
        )
        validate_uploaded_report(other_report, xml_file(month="03"))

        response = self.client.get(reverse("export_archives"))

        self.assertEqual(response.status_code, 200)
        content = b"".join(response.streaming_content)
        with ZipFile(BytesIO(content)) as bundle:
            self.assertIn("J0900108.zip", bundle.namelist())
            self.assertNotIn("2025_Q2/J0900108.zip", bundle.namelist())
            with bundle.open("J0900108.zip") as form_zip_file:
                with ZipFile(BytesIO(form_zip_file.read())) as form_zip:
                    self.assertIn("20809229-2025-Q1.XML", form_zip.namelist())
                    self.assertIn("20809229-2025-Q2.XML", form_zip.namelist())

    def test_organization_admin_dashboard_is_limited_to_managed_organizations(self):
        self.login_organization_admin()
        other = Organization.objects.create(name="Other", edrpou="99999999", report_type=Organization.ReportType.FULL)
        ExpectedReport.objects.create(organization=other, period=self.period, form=self.form)

        response = self.client.get(reverse("admin_dashboard"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "20809229")
        self.assertNotContains(response, "99999999")

    def test_organization_admin_cannot_accept_unmanaged_report(self):
        self.login_organization_admin()
        other = Organization.objects.create(name="Other", edrpou="99999999", report_type=Organization.ReportType.FULL)
        other_report = ExpectedReport.objects.create(organization=other, period=self.period, form=self.form)

        response = self.client.post(reverse("accept_expected_report", args=[other_report.id]))
        other_report.refresh_from_db()

        self.assertEqual(response.status_code, 404)
        self.assertEqual(other_report.status, ExpectedReport.Status.PENDING)

    def test_organization_admin_generation_defaults_to_managed_organizations(self):
        self.login_organization_admin()
        Organization.objects.create(name="Other", edrpou="99999999", report_type=Organization.ReportType.FULL)

        response = self.client.post(
            reverse("generate_expected_reports"),
            {"year": 2026, "quarter": "Q3"},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        period = ReportingPeriod.objects.get(year=2026, quarter="Q3")
        self.assertEqual(ExpectedReport.objects.filter(period=period, organization=self.organization).count(), 2)
        self.assertFalse(ExpectedReport.objects.filter(period=period, organization__edrpou="99999999").exists())

    def test_representative_cannot_open_management_dashboard(self):
        user = get_user_model().objects.create_user(username="representative", password="pass")
        OrganizationUser.objects.create(user=user, organization=self.organization)
        self.client.force_login(user)

        response = self.client.get(reverse("admin_dashboard"))

        self.assertEqual(response.status_code, 302)
