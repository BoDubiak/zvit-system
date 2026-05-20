from openpyxl import Workbook

from .models import ExpectedReport


def build_control_report(period, organization_ids=None):
    reports = (
        ExpectedReport.objects.select_related("organization", "period", "form", "uploaded_by")
        .filter(period=period)
        .order_by("organization__name", "form__code")
    )
    if organization_ids is not None:
        reports = reports.filter(organization_id__in=organization_ids)

    workbook = Workbook()
    received = workbook.active
    received.title = "received_files"
    received.append([
        "organization_name",
        "edrpou",
        "year",
        "quarter",
        "form_code",
        "xml_schema",
        "status",
        "original_filename",
        "normalized_filename",
        "uploaded_at",
        "uploaded_by",
    ])
    for report in reports.exclude(status=ExpectedReport.Status.PENDING):
        received.append([
            report.organization.name,
            report.organization.edrpou,
            report.period.year,
            report.period.quarter,
            report.form.code,
            report.form.xml_schema,
            report.status,
            report.original_filename,
            report.normalized_filename,
            report.uploaded_at.isoformat() if report.uploaded_at else "",
            report.uploaded_by.get_username() if report.uploaded_by else "",
        ])

    missing = workbook.create_sheet("missing_reports")
    missing.append(["organization_name", "edrpou", "year", "quarter", "form_code", "xml_schema", "status"])
    for report in reports.filter(status__in=[ExpectedReport.Status.PENDING, ExpectedReport.Status.ERROR, ExpectedReport.Status.REJECTED]):
        missing.append([
            report.organization.name,
            report.organization.edrpou,
            report.period.year,
            report.period.quarter,
            report.form.code,
            report.form.xml_schema,
            report.status,
        ])

    duplicates = workbook.create_sheet("duplicates")
    duplicates.append(["organization_name", "edrpou", "year", "quarter", "form_code", "xml_schema", "uploads_count"])
    for report in reports:
        count = report.upload_logs.count()
        if count > 1:
            duplicates.append([
                report.organization.name,
                report.organization.edrpou,
                report.period.year,
                report.period.quarter,
                report.form.code,
                report.form.xml_schema,
                count,
            ])

    summary = workbook.create_sheet("summary")
    summary.append(["metric", "value"])
    summary_rows = {
        "organizations_count": reports.values("organization").distinct().count(),
        "expected_reports_count": reports.count(),
        "uploaded_count": reports.filter(status=ExpectedReport.Status.UPLOADED).count(),
        "accepted_count": reports.filter(status=ExpectedReport.Status.ACCEPTED).count(),
        "pending_count": reports.filter(status=ExpectedReport.Status.PENDING).count(),
        "error_count": reports.filter(status=ExpectedReport.Status.ERROR).count(),
        "rejected_count": reports.filter(status=ExpectedReport.Status.REJECTED).count(),
    }
    for metric, value in summary_rows.items():
        summary.append([metric, value])

    return workbook
